import re
import requests
import os
import openpyxl
import exrex

def parse_rules(path_to_file, path_to_file_xlsx):
    dictRules = {}
    ids_in_xlsx = get_ids_in_xlsx(path_to_file_xlsx)
    for conf_file in os.listdir(path_to_file):
        if conf_file.endswith('.conf'):
            with open(path_to_file + '/' + conf_file, encoding='utf-8') as file:
                line = file.readline()

                rule = ''
                chain = False
                start = True

                while line:
                    if not re.search(r'^#|^\s*\n', line):
                        if 'id:' in line:
                            if str(get_id(line)) == '942310':
                                f = 0
                        if 'SecRule' in line and start:
                            rule = line
                            start = False
                        elif 'SecRule' in line and not chain:
                            rule = re.sub(r'\n|\t|\r|(,\\)|(\s\\)', ' ', rule)
                            rule = re.sub(r'\s+', ' ', rule)
                            if get_id(rule) in ids_in_xlsx:
                                part_of_rule = get_part_of_rule(rule)
                                if part_of_rule:
                                    dictRules[get_id(rule)] = get_part_of_rule(rule)
                            rule = line
                        elif re.search(r'\Wchain\W', line):
                            chain = True
                            rule = rule + line
                        elif 'SecRule' in line and chain:
                            chain = False
                            rule = rule + line
                        else:
                            rule = rule + line

                    line = file.readline()
    print('**************************************************************************\n')
    print('parse_rules OK')
    print('\n**************************************************************************')
    return dictRules

def get_ids_in_xlsx(path_to_file_xlsx):
    rules_book = openpyxl.load_workbook(path_to_file_xlsx)
    rules_sheet = rules_book[rules_book.sheetnames[0]]

    list_ids = []
    for row in rules_sheet.iter_rows():
        if row[2].value and row[0].value:
            list_ids.append(str(row[2].value))
    return list_ids

# id = {
#     0: {
#     'args': 'arg1, arg2',
#     'regex': = '',
#     'regex_data': []
#     },
#     1: {
#     'args': 'arg1, arg2',
#     'regex': = '',
#     'regex_data': []
#     }
# }
def get_part_of_rule(rule):
    part_of_rule = {}
    allBlocks = re.findall(r'SecRule\s\S+\s"[^"]+"', re.sub(r'\\"', 'qqqqq', rule))
    count = 0
    for block in allBlocks:
        args = re.search(r'SecRule\s\S+\s', block)
        if not args: raise Exception('Variables not found')
        args = (args.group(0)).replace('SecRule', '').strip()

        regex = re.search(r'"[^"]+"', block).group(0)
        regex = regex.strip('"')
        regex = regex.replace('qqqqq', '\\"')
        operator = re.search(r'@\S+', regex)
        if not operator or operator.group(0) == '@rx':
            regex = regex.replace('@rx ', '')
            regex_data = get_regex_data(regex, 3)
            part_of_rule[count] = {'args': args, 'regex': regex, 'regex_data': regex_data}
            count += 1

    return part_of_rule

def get_id(rule):
    try:
        id = re.search(r'id:\s*\d{6}', rule).group(0)
        id = id.replace('id:', '')
        id = id.strip()
    except:
        raise Exception('ID not found')
    return id

def get_regex_data(regular, num):
    regex_data = []
    try:
        count = exrex.count(regular)  # узнать какое количество вариантов есть
        if count > num: # если возможных вариантов больше чем заданное ограничение, то собираем строки рандомно
            for count in range(num):
                regex_data.append(exrex.getone(regular))
        else: # если возможных вариантов меньше или равно чем заданное ограничение, то собираем все по порядку
            rex = exrex.generate(regular)
            limit = 0
            for line in rex:
                regex_data.append(line)
                if count == limit: break
                count += 1
    except:
        print('Error: ' + regular)
    return regex_data

def get_list_of_successful_regex(path_to_file, path_to_file_xlsx):
    rules = parse_rules(path_to_file, path_to_file_xlsx)
    logs = []
    list_of_successful_regex = {}

    for id, rule in rules.items():
        if len(rule) == 1:
            if len(rule[0]['regex_data']) > 0:
                res = []
                if 'ARGS' in rule[0]['args']:
                    res = send_request(rule[0]['regex_data'], id, args_name='ring')
                elif 'REQUEST_HEADERS' in rule[0]['args']:
                    try:
                        if 'REQUEST_HEADERS:' in rule[0]['args']:
                            header = re.search(r'REQUEST_HEADERS:[^\|\s"]+', rule[0]['args']).group(0)
                            header = (header.replace('REQUEST_HEADERS:', ''))
                            res = send_request(rule[0]['regex_data'], id, header_name=header)
                        else:
                            res = send_request(rule[0]['regex_data'], id, header_name='Host')
                    except:
                        res = []
                elif 'REQUEST_COOKIES' in rule[0]['args']:
                    res = send_request(rule[0]['regex_data'], id, cookies_name='cookiesValue')
                else:
                    logs.append(rule)

                if res:
                    str_of_successful_regex = ''
                    count = 0
                    for regex_data_query_result in res:
                        if regex_data_query_result['status']:
                            str_of_successful_regex = str_of_successful_regex + 'RX' + str(count) + ':' + regex_data_query_result['value'] + '\n'
                        count += 1
                    if str_of_successful_regex:
                        list_of_successful_regex[id] = str_of_successful_regex
        print('****************** End One Rule ******************\n')

    print('****************** Logs in get_list_of_successful_regex ******************\n')
    print(logs)
    print('\n**************************************************************************')
    return list_of_successful_regex



def send_request(values_regex, id, type = 'GET', args_name = '', header_name = '', cookies_name = '', ):
    url = 'http://localhost:8080/'
    path_modsec_audit_in_docker = 'mds:/var/log/modsec_audit.log'
    path_modsec_audit_in_system = 'F:/DockerMDS/test/modsec_audit.log'
    # path_modsec_audit_clean_in_system = 'F:/DockerMDS/test/clean.log'

    result = []

    for value in values_regex:
        # os.system('docker cp ' + path_modsec_audit_clean_in_system + ' ' + path_modsec_audit_in_docker)
        args = {}
        header = {}
        cookies = {}
        if args_name != '':
            args = {args_name: value}
        if header_name != '':
            header = {header_name: value}
        if cookies_name != '':
            cookies = {cookies_name: value}

        try:
            if type == 'OPTIONS':
                response = requests.options(url, params=args, headers=header, cookies=cookies)
            elif type == 'HEAD':
                response = requests.head(url, params=args, headers=header, cookies=cookies)
            elif type == 'POST':
                response = requests.post(url, params=args, headers=header, cookies=cookies)
            elif type == 'PUT':
                response = requests.put(url, params=args, headers=header, cookies=cookies)
            elif type == 'PATCH':
                response = requests.patch(url, params=args, headers=header, cookies=cookies)
            elif type == 'DELETE':
                response = requests.delete(url, params=args, headers=header, cookies=cookies)
            else:
                response = requests.get(url, params=args, headers=header, cookies=cookies)
        except:
            return []
        # response = requests.get(url, headers={'User-Agent': 'Nikto'})

        os.system('docker cp ' + path_modsec_audit_in_docker + ' ' + path_modsec_audit_in_system)
        str_hash = ''
        text_mds = ''

        try:
            with open(path_modsec_audit_in_system, 'r', encoding='utf-8') as f:
                file_text = f.read()
                str_hash = re.search(r'---\S+---Z--\n\n$', file_text).group(0)
                str_hash = str_hash.replace('Z--\n\n', '')
                text_mds = re.search(str_hash + 'H--' + '[\S\s]+' + str_hash + 'I--', file_text).group(0)
        except:
            print(id)

        str_id = 'id "' + id + '"'
        if str_id in text_mds:
            result.append({'value': value, 'status': True})
        else:
            result.append({'value': value, 'status': False})

    return result

def get_xlsx_file(file_name, list_of_successful_regex):
    rules_book = openpyxl.load_workbook(file_name)
    rules_sheet = rules_book[rules_book.sheetnames[0]]

    list_value = []
    for row in rules_sheet.iter_rows():
        rule_mds_id = row[2].value
        if str(rule_mds_id) in list_of_successful_regex:
            list_value.append(list_of_successful_regex[str(rule_mds_id)])
        else:
            list_value.append('')

    count = 1
    for value in list_value:
        try:
            rules_sheet.cell(row=count, column=5).value = value
        except:
            rules_sheet.cell(row=count, column=5).value = 'IllegalCharacterError'
            print(str(count) + ' ' + value)
        count += 1

    rules_book.save(file_name)